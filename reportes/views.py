from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.utils import timezone
from datetime import timedelta
from documentos.models import Documento
from cobros.models import Cobro
from cobradores.models import Cobrador
from openpyxl import Workbook
from django.utils.dateparse import parse_date
import json  # ✅ Añade esta línea junto a las otras importaciones arriba del todo
from django.core.serializers.json import DjangoJSONEncoder  # ✅ Y esta también
from django.core.paginator import Paginator
from clientes.models import Cliente



def localtime_peru():
    return timezone.localtime(timezone.now())


def reporte_clientes_vencidos(request):
    """Clientes con documentos vencidos y saldo pendiente"""
    hoy = localtime_peru().date()
    dias_filtro = request.GET.get('dias', '')  # '30', '60', '90'

    # Base: documentos vencidos con saldo pendiente
    docs_vencidos = Documento.objects.filter(
        fecha_vencimiento__lt=timezone.now(),
        monto_total__gt=F('monto_pagado') + F('monto_devolucion')
    ).select_related('cliente').order_by('fecha_vencimiento')

    # Filtro por rango de días de retraso
    if dias_filtro == '30':
        hace_30 = hoy - timedelta(days=30)
        docs_vencidos = docs_vencidos.filter(fecha_vencimiento__date__gte=hace_30)
    elif dias_filtro == '60':
        hace_30 = hoy - timedelta(days=30)
        hace_60 = hoy - timedelta(days=60)
        docs_vencidos = docs_vencidos.filter(
            fecha_vencimiento__date__lt=hace_30,
            fecha_vencimiento__date__gte=hace_60
        )
    elif dias_filtro == '90':
        hace_60 = hoy - timedelta(days=60)
        docs_vencidos = docs_vencidos.filter(fecha_vencimiento__date__lt=hace_60)

    # Agrupar por cliente
    clientes_data = {}
    for doc in docs_vencidos:
        cliente = doc.cliente
        if cliente.pk not in clientes_data:
            clientes_data[cliente.pk] = {
                'cliente': cliente,
                'documentos': [],
                'total_vencido': 0,
                'dias_promedio': 0,
                'documentos_count': 0
            }
        saldo = doc.get_saldo_pendiente()
        dias_retraso = (hoy - doc.fecha_vencimiento.date()).days
        clientes_data[cliente.pk]['documentos'].append({
            'doc': doc,
            'saldo': saldo,
            'dias_retraso': dias_retraso
        })
        clientes_data[cliente.pk]['total_vencido'] += saldo
        clientes_data[cliente.pk]['dias_promedio'] += dias_retraso
        clientes_data[cliente.pk]['documentos_count'] += 1

    # Calcular promedio
    for data in clientes_data.values():
        if data['documentos_count'] > 0:
            data['dias_promedio'] = data['dias_promedio'] // data['documentos_count']

    clientes_list = sorted(clientes_data.values(), key=lambda x: x['total_vencido'], reverse=True)

    return render(request, 'reportes/clientes_vencidos.html', {
        'clientes_list': clientes_list,
        'dias_filtro': dias_filtro
    })







def reporte_documentos_proximos_vencer(request):
    """Documentos que vencen en los próximos 7 días"""
    hoy = timezone.now().date()
    dentro_de_7_dias = hoy + timedelta(days=7)

    docs = Documento.objects.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=dentro_de_7_dias,
        monto_total__gt=F('monto_pagado') + F('monto_devolucion')  # con saldo pendiente
    ).select_related('cliente', 'cobrador').order_by('fecha_vencimiento')

    # Calcular días restantes
    for doc in docs:
        doc.dias_restantes = (doc.fecha_vencimiento.date() - hoy).days  # ✅ Corregido

    return render(request, 'reportes/documentos_proximos_vencer.html', {
        'docs': docs,
        'hoy': hoy,
        'dentro_de_7_dias': dentro_de_7_dias,
    })

def reporte_top_clientes_pendientes(request):
    """Top 10 clientes con mayor saldo pendiente"""
    hoy = timezone.now().date()

    # Obtener todos los clientes con documentos pendientes
    clientes_data = []
    for cliente in Cliente.objects.all():
        documentos = Documento.objects.filter(
            cliente=cliente,
            monto_total__gt=F('monto_pagado') + F('monto_devolucion')
        )
        if documentos.exists():
            total_pendiente = sum(doc.get_saldo_pendiente() for doc in documentos)
            vencidos = documentos.filter(fecha_vencimiento__lt=hoy).count()
            clientes_data.append({
                'cliente': cliente,
                'total_pendiente': total_pendiente,
                'documentos_pendientes': documentos.count(),
                'documentos_vencidos': vencidos,
            })

    # Ordenar por saldo pendiente y tomar top 10
    clientes_data.sort(key=lambda x: x['total_pendiente'], reverse=True)
    top_clientes = clientes_data[:10]

    return render(request, 'reportes/top_clientes_pendientes.html', {
        'top_clientes': top_clientes,
        'hoy': hoy,
    })









def reporte_cobradores(request):
    """Monto total cobrado por cada cobrador (filtrado por fecha)"""
    # Filtro de fecha
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    filtro_rapido = request.GET.get('filtro')  # 'hoy', 'ayer', 'mes', etc.

    hoy = localtime_peru().date()

    # Definir rangos según el filtro rápido
    if filtro_rapido == 'hoy':
        fecha_desde = fecha_hasta = hoy.isoformat()
    elif filtro_rapido == 'ayer':
        ayer = hoy - timedelta(days=1)
        fecha_desde = fecha_hasta = ayer.isoformat()
    elif filtro_rapido == 'mes':
        fecha_desde = (hoy.replace(day=1)).isoformat()
        fecha_hasta = hoy.isoformat()
    elif filtro_rapido == 'mes_pasado':
        primer_dia = (hoy.replace(day=1) - timedelta(days=1)).replace(day=1)
        ultimo_dia = hoy.replace(day=1) - timedelta(days=1)
        fecha_desde = primer_dia.isoformat()
        fecha_hasta = ultimo_dia.isoformat()
    elif filtro_rapido == '3meses':
        fecha_desde = (hoy - timedelta(days=90)).isoformat()
        fecha_hasta = hoy.isoformat()
    elif filtro_rapido == 'año':
        fecha_desde = hoy.replace(month=1, day=1).isoformat()
        fecha_hasta = hoy.isoformat()
    elif filtro_rapido == 'año_pasado':
        año_pasado = hoy.year - 1
        fecha_desde = f"{año_pasado}-01-01"
        fecha_hasta = f"{año_pasado}-12-31"

    # Valores por defecto si no hay filtros
    if not fecha_desde:
        fecha_desde = (hoy - timedelta(days=30)).isoformat()
    if not fecha_hasta:
        fecha_hasta = hoy.isoformat()

    # Convertir a datetime
    from django.utils.dateparse import parse_date
    fecha_desde_dt = parse_date(fecha_desde)
    fecha_hasta_dt = parse_date(fecha_hasta)

    if not fecha_desde_dt:
        fecha_desde_dt = hoy - timedelta(days=30)
    if not fecha_hasta_dt:
        fecha_hasta_dt = hoy

    # Filtrar cobros
    cobros = Cobro.objects.filter(
        fecha__date__gte=fecha_desde_dt,
        fecha__date__lte=fecha_hasta_dt
    ).select_related('cobrador')

    # Agrupar por cobrador
    from collections import defaultdict
    total_por_cobrador = defaultdict(float)
    for cobro in cobros:
        total_por_cobrador[cobro.cobrador] += float(cobro.monto)

    cobradores_data = [
        {'cobrador': c, 'total_cobrado': t}
        for c, t in total_por_cobrador.items()
    ]
    cobradores_data.sort(key=lambda x: x['total_cobrado'], reverse=True)
    total_general = sum(d['total_cobrado'] for d in cobradores_data)

    # ✅ Convertir datos a JSON para el gráfico
    cobradores_json = json.dumps([
        {'nombre': d['cobrador'].nombre, 'total': d['total_cobrado']}
        for d in cobradores_data
    ], cls=DjangoJSONEncoder)

    return render(request, 'reportes/cobradores.html', {
        'cobradores_data': cobradores_data,
        'total_general': total_general,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'filtro_rapido': filtro_rapido,
        'cobradores_json': cobradores_json,  # ✅ Pasamos los datos al template
    })


def reporte_pagos_parciales(request):
    """Documentos con pagos parciales (no pagados ni sin pago)"""
    docs = Documento.objects.filter(
        monto_pagado__gt=0,
        monto_pagado__lt=F('monto_total') - F('monto_devolucion')
    ).select_related('cliente').order_by('-monto_pagado')

    # Calcular datos adicionales
    docs_data = []
    total_documentos = 0
    total_facturado = 0
    total_pagado = 0
    total_saldo = 0

    for doc in docs:
        # ✅ Usa los campos directamente del documento
        pagado = doc.monto_pagado
        devolucion = doc.monto_devolucion
        total = doc.monto_total
        saldo = doc.get_saldo_pendiente()

        # ✅ Evita división por cero y limita a 100%
        if total > 0:
            porcentaje = min(int((pagado / total) * 100), 100)
        else:
            porcentaje = 0
        print(f"Documento: {doc}, Pagado: {pagado}, Total: {total}, Porcentaje: {porcentaje}")
        docs_data.append({
            'doc': doc,
            'pagado': pagado,
            'devolucion': devolucion,
            'total': total,
            'saldo': saldo,
            'porcentaje': porcentaje  # ✅ Ahora es un número entre 0 y 100
        })

        # Acumuladores
        total_documentos += 1
        total_facturado += total
        total_pagado += pagado
        total_saldo += saldo

    # ✅ Paginación: 3 documentos por página
    paginator = Paginator(docs_data, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'reportes/pagos_parciales.html', {
        'page_obj': page_obj,
        'total_documentos': total_documentos,
        'total_facturado': total_facturado,
        'total_pagado': total_pagado,
        'total_saldo': total_saldo,
    })


def reporte_totales(request):
    """Totales generales del sistema"""
    # Todos los documentos
    docs = Documento.objects.all()

    # Totales
    total_facturado = docs.aggregate(total=Sum('monto_total'))['total'] or 0
    total_pagado = docs.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_devolucion = docs.aggregate(total=Sum('monto_devolucion'))['total'] or 0
    total_pendiente = total_facturado - total_pagado - total_devolucion

    # Documentos vencidos con saldo pendiente
    vencidos = docs.filter(
        fecha_vencimiento__lt=timezone.now(),
        monto_total__gt=F('monto_pagado') + F('monto_devolucion')
    )
    total_vencido = sum(doc.get_saldo_pendiente() for doc in vencidos)

    # ✅ Conteo por tipo de documento
    count_facturas = docs.filter(tipo='factura').count()
    count_boletas = docs.filter(tipo='boleta').count()
    count_notas = docs.filter(tipo='nota_venta').count()

    return render(request, 'reportes/totales.html', {
        'total_facturado': total_facturado,
        'total_pagado': total_pagado,
        'total_devolucion': total_devolucion,
        'total_pendiente': total_pendiente,
        'total_vencido': total_vencido,
        'count_documentos': docs.count(),
        'count_clientes': docs.values('cliente').distinct().count(),
        # ✅ Añadido: conteo por tipo
        'count_facturas': count_facturas,
        'count_boletas': count_boletas,
        'count_notas': count_notas,
    })


# Exportación a Excel
def exportar_excel(request, tipo):
    """Exporta cualquier reporte a Excel (con filtros)"""
    workbook = Workbook()

    if tipo == 'clientes_vencidos':
        return _exportar_clientes_vencidos(workbook)
    elif tipo == 'cobradores':
        return _exportar_cobradores(workbook, request)  # ✅ Pasamos el request
    elif tipo == 'pagos_parciales':
        return _exportar_pagos_parciales(workbook)
    elif tipo == 'totales':
        return _exportar_totales(workbook)

    # Respuesta por defecto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte.xlsx'
    workbook.save(response)
    return response


def _exportar_clientes_vencidos(workbook):
    sheet = workbook.active
    sheet.title = "Clientes Vencidos"

    headers = ['Cliente', 'DNI/RUC', 'Documento', 'Monto Total', 'Pagado', 'Devolución', 'Saldo', 'Vencimiento', 'Días Retraso']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    hoy = localtime_peru().date()
    docs = Documento.objects.filter(
        fecha_vencimiento__lt=timezone.now(),
        monto_total__gt=F('monto_pagado') + F('monto_devolucion')
    ).select_related('cliente')

    for doc in docs:
        saldo = doc.get_saldo_pendiente()
        dias = (hoy - doc.fecha_vencimiento.date()).days
        sheet.append([
            doc.cliente.nombre,
            doc.cliente.dni_ruc,
            f"{doc.get_tipo_display()} {doc.get_numero_completo()}",
            float(doc.monto_total),
            float(doc.monto_pagado),
            float(doc.monto_devolucion),
            float(saldo),
            doc.fecha_vencimiento.strftime('%d/%m/%Y'),
            dias
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes_vencidos.xlsx'
    workbook.save(response)
    return response


def _exportar_cobradores(workbook, request=None):
    """
    Exporta cobradores a Excel (con filtros si se pasan)
    request=None para mantener compatibilidad
    """
    sheet = workbook.active
    sheet.title = "Cobradores"

    headers = ['Cobrador', 'DNI', 'Total Cobrado']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    # Obtener fechas del request si existen
    if request:
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        fecha_desde_dt = parse_date(fecha_desde) if fecha_desde else None
        fecha_hasta_dt = parse_date(fecha_hasta) if fecha_hasta else None
    else:
        fecha_desde_dt = fecha_hasta_dt = None

    # Aplicar filtros
    cobros = Cobro.objects.all()
    if fecha_desde_dt:
        cobros = cobros.filter(fecha__date__gte=fecha_desde_dt)
    if fecha_hasta_dt:
        cobros = cobros.filter(fecha__date__lte=fecha_hasta_dt)

    # Agrupar
    from collections import defaultdict
    total_por_cobrador = defaultdict(float)
    for cobro in cobros.select_related('cobrador'):
        total_por_cobrador[cobro.cobrador] += float(cobro.monto)

    # Agregar filas
    for cobrador, total in total_por_cobrador.items():
        sheet.append([cobrador.nombre, cobrador.dni, float(total)])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cobradores.xlsx'
    workbook.save(response)
    return response


def _exportar_pagos_parciales(workbook):
    sheet = workbook.active
    sheet.title = "Pagos Parciales"

    headers = ['Cliente', 'Documento', 'Total', 'Pagado', 'Devolución', 'Saldo', 'Porcentaje']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    docs = Documento.objects.filter(
        monto_pagado__gt=0,
        monto_pagado__lt=F('monto_total') - F('monto_devolucion')
    ).select_related('cliente')

    for doc in docs:
        pagado = doc.monto_pagado
        dev = doc.monto_devolucion
        total = doc.monto_total
        saldo = total - pagado - dev
        porcentaje = (pagado / total * 100) if total > 0 else 0
        sheet.append([
            doc.cliente.nombre,
            f"{doc.get_tipo_display()} {doc.get_numero_completo()}",
            float(total),
            float(pagado),
            float(dev),
            float(saldo),
            f"{porcentaje:.2f}%"
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=parciales.xlsx'
    workbook.save(response)
    return response


def _exportar_totales(workbook):
    sheet = workbook.active
    sheet.title = "Totales"

    # Datos
    docs = Documento.objects.all()
    total_facturado = docs.aggregate(t=Sum('monto_total'))['t'] or 0
    total_pagado = docs.aggregate(t=Sum('monto_pagado'))['t'] or 0
    total_devolucion = docs.aggregate(t=Sum('monto_devolucion'))['t'] or 0
    total_pendiente = total_facturado - total_pagado - total_devolucion

    data = [
        ["Concepto", "Monto (S/)"],
        ["Total Facturado", float(total_facturado)],
        ["Total Pagado", float(total_pagado)],
        ["Total Devolución", float(total_devolucion)],
        ["Total Pendiente", float(total_pendiente)],
        ["Clientes Únicos", docs.values('cliente').distinct().count()],
        ["Documentos", docs.count()],
    ]

    for row in data:
        sheet.append(row)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=totales.xlsx'
    workbook.save(response)
    return response