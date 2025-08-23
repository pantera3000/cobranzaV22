from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum  # ✅ Asegúrate de tener Sum
from django.db.models import F
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from clientes.models import Cliente
from django.http import HttpResponseRedirect
import csv
from openpyxl import Workbook
from .models import Cobro
from .forms import CobroForm
from documentos.models import Documento
import json
from datetime import date
import calendar
from cobradores.models import Cobrador  # ✅ Falta esta línea
from decimal import Decimal
from openpyxl import Workbook
from django.http import HttpResponse
from calendar import monthrange
from datetime import datetime

from clientes.utils import registrar_log  # ✅ Importar
from django.contrib.auth.decorators import login_required
from django.db import transaction
import json
from django.contrib.auth.decorators import permission_required
from clientes.models import LogActividad
from django.db.models import Count, Sum, Max  # ✅ Usamos agregaciones de Django




@permission_required('clientes.view_logactividad', raise_exception=True)
def log_actividad(request):
    logs = LogActividad.objects.all().select_related('usuario', 'cobrador').order_by('-fecha')
    
    categoria = request.GET.get('categoria')
    if categoria:
        logs = logs.filter(categoria=categoria)

    paginator = Paginator(logs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'clientes/log_actividad.html', {
        'page_obj': page_obj,
        'categoria': categoria,
        'categorias': LogActividad.CATEGORIA_OPCIONES
    })




@transaction.atomic
def cobro_create(request):
    documento_inicial = None
    documento_inicial_data = None

    if request.method == 'POST':
        form = CobroForm(request.POST)
        if form.is_valid():
            cobro = form.save(commit=False)
            documento = cobro.documento

            saldo_pendiente = documento.get_saldo_pendiente()
            if cobro.monto > saldo_pendiente:
                messages.error(
                    request,
                    f"El monto no puede exceder el saldo pendiente de S/ {saldo_pendiente:,.2f}."
                )
            else:
                cobro.save()

                # ✅ Registrar log
                registrar_log(
                    usuario=request.user,  # 👈 El usuario logueado
                    cobrador=cobro.cobrador,
                    categoria='cobro',
                    accion='Registró pago',
                    descripcion=f"Monto: S/ {cobro.monto:,.2f}, Documento: {documento}, Cliente: {documento.cliente.nombre}, Referencia: {cobro.referencia or '-'}"
                )

                messages.success(
                    request,
                    f"Pago de S/ {cobro.monto:,.2f} registrado exitosamente para {documento}."
                )

                next_url = request.POST.get('next')
                if next_url:
                    return redirect(next_url)
                return redirect('cobros:cobro_list')
        else:
            messages.error(request, 'Por favor corrige los errores.')
    else:
        documento_id = request.GET.get('documento')
        initial = {}

        if documento_id:
            try:
                doc = Documento.objects.get(pk=documento_id)
                saldo = doc.get_saldo_pendiente()

                if saldo > 0:
                    initial['documento'] = doc
                    documento_inicial = doc

                    data_dict = {
                        'id': doc.id,
                        'tipo_display': str(doc.get_tipo_display()),
                        'numero_completo': str(doc.get_numero_completo()),
                        'cliente_nombre': str(doc.cliente.nombre),
                        'cliente_dni': str(doc.cliente.dni_ruc),
                        'monto_total': float(doc.monto_total),
                        'saldo_pendiente': float(saldo),
                    }
                    documento_inicial_data = json.dumps(data_dict)
            except Documento.DoesNotExist:
                pass

        form = CobroForm(initial=initial)

    return render(request, 'cobros/cobro_form.html', {
        'form': form,
        'title': 'Registrar Pago',
        'documento_inicial': documento_inicial,
        'documento_inicial_data': documento_inicial_data,
    })

from django.db.models import Q, Sum  # ✅ Asegúrate de tener Sum

def cobro_list(request):
    query = request.GET.get('q', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    cobrador_id = request.GET.get('cobrador', '')

    # Iniciar queryset
    cobros = Cobro.objects.select_related('documento', 'documento__cliente', 'cobrador').all()

    # Filtros
    if query:
        cobros = cobros.filter(
            Q(documento__numero__icontains=query) |
            Q(documento__cliente__nombre__icontains=query) |
            Q(documento__cliente__dni_ruc__icontains=query) |
            Q(cobrador__nombre__icontains=query)
        )

    if fecha_desde:
        cobros = cobros.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        cobros = cobros.filter(fecha__date__lte=fecha_hasta)

    if cobrador_id:
        try:
            cobros = cobros.filter(cobrador_id=int(cobrador_id))
        except (ValueError, TypeError):
            cobrador_id = None  # Si no es válido, ignora el filtro

    # Obtener el cobrador seleccionado para mostrar en el filtro activo
    cobrador_seleccionado = None
    if cobrador_id:
        try:
            cobrador_seleccionado = Cobrador.objects.get(id=cobrador_id)
        except Cobrador.DoesNotExist:
            cobrador_id = None  # Si no existe, ignora

    # Calcular total cobrado
    total_cobrado = cobros.aggregate(total=Sum('monto'))['total'] or 0

    # Paginación
    paginator = Paginator(cobros, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Opciones para filtros rápidos
    cobradores = Cobrador.objects.all().order_by('nombre')
    today = date.today()
    mes_inicio = today.replace(day=1)
    mes_fin = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    año_actual = today.year
    año_pasado = today.year - 1

    # Mes pasado
    if today.month == 1:
        mes_pasado_inicio = today.replace(year=today.year - 1, month=12, day=1)
        mes_pasado_fin = today.replace(year=today.year - 1, month=12, day=31)
    else:
        mes_pasado_inicio = today.replace(month=today.month - 1, day=1)
        mes_pasado_fin = today.replace(month=today.month - 1, day=calendar.monthrange(today.year, today.month - 1)[1])

    # Contexto completo
    context = {
        'page_obj': page_obj,
        'query': query,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'cobrador_id': cobrador_id,
        'cobrador_seleccionado': cobrador_seleccionado,
        'cobradores': cobradores,
        'today': today,
        'mes_inicio': mes_inicio,
        'mes_fin': mes_fin,
        'mes_pasado_inicio': mes_pasado_inicio,
        'mes_pasado_fin': mes_pasado_fin,
        'total_cobrado': total_cobrado,
        'año_actual': año_actual,
        'año_pasado': año_pasado,
    }

    return render(request, 'cobros/cobro_list.html', context)




    return render(request, 'cobros/cobro_list.html', {
        'page_obj': page_obj,
        'query': query,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    })


# cobros/views.py
def cobro_export_excel(request):
    """
    Exporta los cobros a Excel, aplicando los mismos filtros que en la vista de listado.
    Incluye las columnas 'Referencia' y 'Notas'.
    """
    # === 1. Obtener filtros (igual que en cobro_list) ===
    query = request.GET.get('q', '')
    cobrador_id = request.GET.get('cobrador')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    # === 2. Obtener y filtrar cobros ===
    cobros = Cobro.objects.select_related(
        'documento', 'documento__cliente', 'cobrador'
    ).all().order_by('-fecha')

    # Filtro por búsqueda
    if query:
        cobros = cobros.filter(
            Q(documento__numero__icontains=query) |
            Q(documento__serie__icontains=query) |
            Q(documento__cliente__nombre__icontains=query) |
            Q(referencia__icontains=query) |
            Q(notas__icontains=query)  # ✅ Opcional: buscar también en notas
        )

    # Filtro por cobrador
    if cobrador_id:
        cobros = cobros.filter(cobrador__id=cobrador_id)

    # Filtro por fecha
    if fecha_desde:
        cobros = cobros.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        cobros = cobros.filter(fecha__date__lte=fecha_hasta)

    # === 3. Crear libro de Excel ===
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cobros"

    # === 4. Encabezados (con Referencia y Notas) ===
    headers = [
        'Documento', 
        'Cliente', 
        'Monto', 
        'Cobrador', 
        'Referencia',
        'Notas',
        'Fecha Pago', 
        'Fecha Registro'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header

    # === 5. Agregar datos ===
    for cobro in cobros:
        sheet.append([
            f"{cobro.documento.get_tipo_display()} {cobro.documento.serie}-{cobro.documento.numero}",
            cobro.documento.cliente.nombre,
            float(cobro.monto),
            cobro.cobrador.nombre,
            cobro.referencia or "",
            cobro.notas or "",
            cobro.fecha.strftime('%d/%m/%Y %H:%M'),
            cobro.creado_en.strftime('%d/%m/%Y %H:%M'),
        ])

    # === 6. Ajustar ancho de columnas (opcional, mejora visual) ===
    column_widths = [18, 30, 12, 20, 20, 30, 18, 18]
    for i, width in enumerate(column_widths, 1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # === 7. Preparar respuesta HTTP ===
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cobros.xlsx'
    workbook.save(response)
    return response


def cobro_export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=cobros.csv'

    writer = csv.writer(response)
    writer.writerow(['Documento', 'Cliente', 'Monto', 'Cobrador', 'Fecha Pago', 'Fecha Registro'])

    for cobro in Cobro.objects.select_related('documento', 'documento__cliente', 'cobrador').all():
        writer.writerow([
            f"{cobro.documento.get_tipo_display()} {cobro.documento.serie}-{cobro.documento.numero}",
            cobro.documento.cliente.nombre,
            cobro.monto,
            cobro.cobrador.nombre,
            cobro.fecha.strftime('%d/%m/%Y %H:%M'),
            cobro.creado_en.strftime('%d/%m/%Y %H:%M'),
        ])

    return response




def cobro_delete(request, pk):
    cobro = get_object_or_404(Cobro, pk=pk)
    if request.method == 'POST':
        monto = cobro.monto
        # Guardamos datos antes de eliminar
        documento = cobro.documento
        cliente_nombre = documento.cliente.nombre
        documento_numero = f"{documento.get_tipo_display()} {documento.get_numero_completo()}"
        
        cobro.delete()  # ← Aquí se actualiza monto_pagado (vía modelo)
        
        # ✅ Registrar en el log
        registrar_log(
            usuario=request.user,
            cobrador=cobro.cobrador,
            categoria='cobro',
            accion='Eliminó pago',
            descripcion=f"Monto: S/ {monto:,.2f}, Documento: {documento_numero}, Cliente: {cliente_nombre}"
        )

        messages.success(request, f'Pago de S/ {monto:,.2f} eliminado correctamente.')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'))
    return render(request, 'cobros/cobro_confirm_delete.html', {
        'cobro': cobro
    })









def pago_multiple(request):
    """Página para realizar múltiples pagos a la vez"""
    query = request.GET.get('q', '')
    page_number = request.GET.get('page', 1)

    # Solo buscar si hay consulta
    if query:
        documentos = Documento.objects.filter(
            monto_total__gt=F('monto_pagado') + F('monto_devolucion')
        ).filter(
            Q(cliente__nombre__icontains=query) |
            Q(numero__icontains=query) |
            Q(serie__icontains=query)
        ).select_related('cliente').order_by('-fecha_emision')
    else:
        # ✅ No cargar todos si no hay búsqueda
        # Mostrar solo los 10 más recientes
        documentos = Documento.objects.filter(
            monto_total__gt=F('monto_pagado') + F('monto_devolucion')
        ).select_related('cliente').order_by('-fecha_emision')[:10]

    # ✅ Paginación
    paginator = Paginator(documentos, 3)
    documentos_page = paginator.get_page(page_number)

    cobradores = Cobrador.objects.all()

    return render(request, 'cobros/pago_multiple.html', {
        'documentos': documentos_page,
        'query': query,
        'cobradores': cobradores
    })

# cobros/views.py
def registrar_pagos_multiple(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    try:
        data = request.POST
        print("🔍 POST completo recibido:", dict(data))  # ✅ Añade esta línea

        cobrador_id = data.get('cobrador')
        cobrador = get_object_or_404(Cobrador, pk=cobrador_id)

        # ✅ Obtener referencia
        referencia = data.get('referencia', '').strip()
        notas = data.get('notas', '').strip()  # ✅ Obtener notas

        print(f"✅ Referencia: '{referencia}'")  # ✅ Depuración
        print(f"✅ Notas: '{notas}'")             # ✅ Depuración

        pagos = []
        total_registrado = 0

        for key in data:
            if key.startswith('pago_'):
                doc_id = key.replace('pago_', '')
                monto_str = data[key].replace(',', '.')
                try:
                    monto = float(monto_str)
                    if monto <= 0:
                        continue
                except ValueError:
                    continue

                documento = get_object_or_404(Documento, pk=doc_id)
                saldo = documento.get_saldo_pendiente()
                if monto > saldo:
                    return JsonResponse({
                        'error': f'El monto para {documento.get_tipo_display()} {documento.get_numero_completo} excede el saldo pendiente.'
                    }, status=400)

                # ✅ Crear cobro con referencia
                cobro = Cobro(
                    documento=documento,
                    cobrador=cobrador,
                    monto=monto,
                    fecha=timezone.now(),
                    referencia=referencia,  # ✅ Guardar referencia
                    notas=notas  # ✅ Guardar notas
                )
                cobro.save()
                total_registrado += monto

        return JsonResponse({
            'success': True,
            'total_registrado': total_registrado,
            'redirect_url': request.META.get('HTTP_REFERER', '/')
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    







def buscar_por_referencia(request):
    """Buscar pagos por referencia"""
    query = request.GET.get('q', '')
    pagos = []
    total_monto = 0

    if query:
        pagos = Cobro.objects.filter(
            referencia__icontains=query
        ).select_related('documento', 'documento__cliente', 'cobrador').order_by('-fecha')
        total_monto = sum(cobro.monto for cobro in pagos)

    return render(request, 'cobros/buscar_por_referencia.html', {
        'query': query,
        'pagos': pagos,
        'total_monto': total_monto
    })


def historial_referencias(request):
    """Listado de todas las referencias con filtros y búsqueda"""
    # Obtener filtros
    filtro_fecha = request.GET.get('fecha')
    query = request.GET.get('q', '').strip()  # ✅ Obtener búsqueda

    # Empezamos con cobros que tengan referencia
    cobros = Cobro.objects.exclude(referencia__isnull=True).exclude(referencia='')

    # Aplicar búsqueda por referencia (si hay query)
    if query:
        cobros = cobros.filter(referencia__icontains=query)

    # Aplicar filtro por fecha usando zona horaria de Lima
    hoy = timezone.localtime(timezone.now()).date()

    if filtro_fecha == 'hoy':
        cobros = cobros.filter(fecha__date=hoy)
    elif filtro_fecha == 'ayer':
        ayer = hoy - timedelta(days=1)
        cobros = cobros.filter(fecha__date=ayer)
    elif filtro_fecha == 'semana':
        inicio_semana = hoy - timedelta(days=hoy.weekday())  # Lunes de esta semana
        cobros = cobros.filter(fecha__date__gte=inicio_semana)
    elif filtro_fecha == 'mes':
        cobros = cobros.filter(fecha__date__year=hoy.year, fecha__date__month=hoy.month)

    # Agrupar por referencia
    referencias = (
        cobros.values('referencia')
        .annotate(
            count=Count('id'),
            total=Sum('monto'),
            fecha=Max('fecha')  # Último pago con esa referencia
        )
        .order_by('-fecha')
    )

    # Paginación
    paginator = Paginator(referencias, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'cobros/historial_referencias.html', {
        'page_obj': page_obj,
        'filtro_fecha': filtro_fecha,
        'query': query,  # ✅ Pasar query al template
    })








def exportar_por_referencia(request):
    """Exportar pagos por referencia a Excel"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        # Si no hay referencia, devolver vacío o error
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos por Referencia"
        ws.append(["Error: No se especificó una referencia"])
        filename = f"pagos_sin_referencia_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:
        # Filtrar pagos por referencia
        pagos = Cobro.objects.filter(
            referencia__icontains=query
        ).select_related('documento', 'documento__cliente', 'cobrador').order_by('-fecha')

        wb = Workbook()
        ws = wb.active
        ws.title = f"Pagos - {query}"

        # Encabezados
        headers = ['Documento', 'Cliente', 'Monto (S/)', 'Cobrador', 'Fecha Pago', 'Referencia']
        ws.append(headers)

        # Datos
        for cobro in pagos:
            ws.append([
                f"{cobro.documento.get_tipo_display()} {cobro.documento.get_numero_completo()}",
                cobro.documento.cliente.nombre,
                float(cobro.monto),
                cobro.cobrador.nombre,
                cobro.fecha.strftime('%d/%m/%Y %H:%M'),
                cobro.referencia or ''
            ])

        # Ajustar ancho de columnas
        for col in ['A', 'B', 'F']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18

        filename = f"pagos_{query}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response






def reporte_cartera(request):
    """Reporte de Cartera de Cobranzas"""
    today = timezone.now().date()

    # === 1. Resumen General ===
    documentos = Documento.objects.all()
    total_facturado = documentos.aggregate(total=Sum('monto_total'))['total'] or 0
    total_pagado = sum(doc.monto_pagado for doc in documentos)
    total_devolucion = sum(doc.monto_devolucion for doc in documentos)
    total_pendiente = total_facturado - total_pagado - total_devolucion

    # Documentos vencidos (con saldo pendiente)
    vencidos = documentos.filter(
        fecha_vencimiento__lt=timezone.now(),
        monto_total__gt=F('monto_pagado') + F('monto_devolucion')
    )
    total_vencido = sum(doc.get_saldo_pendiente() for doc in vencidos)

    # === 2. Top 10 Clientes con Mayor Saldo Pendiente ===
    clientes_pendientes = []
    for cliente in Cliente.objects.all():
        docs = documentos.filter(cliente=cliente)
        saldo = sum(doc.get_saldo_pendiente() for doc in docs)
        if saldo > 0:
            clientes_pendientes.append({
                'cliente': cliente,
                'saldo': saldo,
                'vencido': sum(
                    doc.get_saldo_pendiente()
                    for doc in docs
                    if doc.fecha_vencimiento < timezone.now()
                )
            })
    # Ordenar por saldo descendente
    clientes_pendientes = sorted(clientes_pendientes, key=lambda x: x['saldo'], reverse=True)[:10]

    # === 3. Resumen por Cobrador ===
    cobradores_data = []
    for cobrador in Cobrador.objects.all():
        cobros = Cobro.objects.filter(cobrador=cobrador)
        total = cobros.aggregate(total=Sum('monto'))['total'] or 0
        if total > 0:
            cobradores_data.append({
                'cobrador': cobrador,
                'total_cobrado': total,
            })
    cobradores_data = sorted(cobradores_data, key=lambda x: x['total_cobrado'], reverse=True)

    # === 4. Evolución Mensual (últimos 6 meses) ===
    evolucion = []
    for i in range(6):
        mes = today - timedelta(days=30 * i)
        primer_dia = mes.replace(day=1)
        _, last_day = monthrange(mes.year, mes.month)
        ultimo_dia = mes.replace(day=last_day)

        fecha_desde_dt = timezone.make_aware(datetime.combine(primer_dia, datetime.min.time()))
        fecha_hasta_dt = timezone.make_aware(datetime.combine(ultimo_dia, datetime.max.time()))

        cobros_mes = Cobro.objects.filter(
            fecha__gte=fecha_desde_dt,
            fecha__lte=fecha_hasta_dt
        ).aggregate(total=Sum('monto'))['total'] or 0

        # ✅ Convertir a float para que sea compatible con JSON
        evolucion.append({
            'mes': primer_dia.strftime('%b %Y'),  # Ej: "Aug 2025"
            'cobrado': float(cobros_mes)  # ✅ Forzar a float
        })
    evolucion.reverse()
    print("📊 Evolución de cobros:", evolucion)  # 👈 Depuración

    context = {
        'total_facturado': total_facturado,
        'total_pagado': total_pagado,
        'total_devolucion': total_devolucion,
        'total_pendiente': total_pendiente,
        'total_vencido': total_vencido,
        'clientes_pendientes': clientes_pendientes,
        'cobradores_data': cobradores_data,
        'evolucion': evolucion,
        'fecha_reporte': today,
    }

    return render(request, 'cobros/reporte_cartera.html', context)









def exportar_cartera_excel(request):
    """Exportar reporte de cartera a Excel"""
    # Reutiliza la lógica de resumen
    documentos = Documento.objects.all()
    total_facturado = documentos.aggregate(total=Sum('monto_total'))['total'] or 0
    total_pagado = sum(doc.monto_pagado for doc in documentos)
    total_devolucion = sum(doc.monto_devolucion for doc in documentos)
    total_pendiente = total_facturado - total_pagado - total_devolucion

    vencidos = documentos.filter(
        fecha_vencimiento__lt=timezone.now(),
        monto_total__gt=F('monto_pagado') + F('monto_devolucion')
    )
    total_vencido = sum(doc.get_saldo_pendiente() for doc in vencidos)

    # Crear libro
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Cartera"

    # Encabezado
    ws.append(["REPORTE DE CARTERA DE COBRANZAS", "", "", "", "", ""])
    ws.append(["Fecha del Reporte:", timezone.now().date().strftime('%d/%m/%Y'), "", "", "", ""])
    ws.append([])
    ws.append(["RESUMEN GENERAL", "", "", "", "", ""])
    ws.append(["Total Facturado", total_facturado])
    ws.append(["Total Cobrado", total_pagado])
    ws.append(["Devoluciones", total_devolucion])
    ws.append(["Saldo Pendiente", total_pendiente])
    ws.append(["Vencido", total_vencido])
    ws.append([])

    # Top clientes
    ws.append(["TOP 10 CLIENTES CON SALDO PENDIENTE", "", ""])
    ws.append(["Cliente", "Saldo", "Vencido"])
    for item in sorted(
        [{'cliente': c, 'saldo': sum(doc.get_saldo_pendiente() for doc in Documento.objects.filter(cliente=c)), 'vencido': sum(doc.get_saldo_pendiente() for doc in Documento.objects.filter(cliente=c, fecha_vencimiento__lt=timezone.now()))} for c in Cliente.objects.all() if sum(doc.get_saldo_pendiente() for doc in Documento.objects.filter(cliente=c)) > 0],
        key=lambda x: x['saldo'], reverse=True
    )[:10]:
        ws.append([item['cliente'].nombre, item['saldo'], item['vencido']])

    # Ajustar ancho
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_cartera.xlsx'
    wb.save(response)
    return response